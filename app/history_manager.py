# -*- coding: utf-8 -*-
# Управление версиями файлов (хранилищем)
import hashlib
import os
import shutil
import sqlite3
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional, Dict
import tempfile 
import psutil # Для получения информации о диске

from PySide6.QtCore import QObject, Signal, Slot, QThread, QTimer
from PySide6.QtWidgets import QSystemTrayIcon

# --- Импорты для обработки файлов ---
from PIL import Image 
import fitz 
from docx import Document 
from openpyxl import load_workbook 


class ScanWorker(QObject):
    """
    Рабочий, выполняющий сканирование файлов в отдельном потоке.
    Теперь понимает сложную структуру 'watched_items' с исключениями.
    """
    finished = Signal()
    progress = Signal(str)
    scan_notification = Signal(str, QSystemTrayIcon.MessageIcon)

    def __init__(self, history_manager, items_to_scan: List[Dict]):
        super().__init__()
        self.history_manager = history_manager
        self.items_to_scan = items_to_scan
        self._should_stop = False

    def stop(self):
        self._should_stop = True

    def run(self):
        self._should_stop = False
        try:
            self.scan_notification.emit(
                self.tr("Началось фоновое сканирование файлов..."),
                QSystemTrayIcon.Information
            )
            for item in self.items_to_scan:
                if self._should_stop: break

                path_str = item.get("path")
                item_type = item.get("type")
                exclusions = item.get("exclusions", [])

                path = Path(path_str)
                if not path.exists():
                    self.scan_notification.emit(
                        self.tr("Путь для сканирования не существует и будет проигнорирован: {0}").format(path_str),
                        QSystemTrayIcon.Warning
                    )
                    continue

                if item_type == "file":
                    self.progress.emit(path.name)
                    try:
                        self.history_manager.add_initial_version(path)
                    except Exception as e:
                        self.scan_notification.emit(self.tr("Ошибка при обработке {0}: {1}").format(path, e), QSystemTrayIcon.Warning)

                elif item_type == "folder":
                    # Преобразуем строки исключений в Path объекты для корректной работы
                    exclusion_paths = {Path(ex).resolve() for ex in exclusions}

                    for root, dirs, files in os.walk(path):
                        if self._should_stop: break

                        # --- Эффективная логика исключений ---
                        # Модифицируем 'dirs' на месте, чтобы os.walk не заходил в исключенные папки
                        dirs[:] = [d for d in dirs if Path(root, d).resolve() not in exclusion_paths]

                        for name in files:
                            if self._should_stop: break
                            file_path = Path(root) / name
                            self.progress.emit(file_path.name)
                            try:
                                self.history_manager.add_initial_version(file_path)
                            except Exception as e:
                                self.scan_notification.emit(self.tr("Ошибка при обработке {0}: {1}").format(file_path, e), QSystemTrayIcon.Warning)
                    if self._should_stop: self.scan_notification.emit(self.tr("Сканирование прервано пользователем."), QSystemTrayIcon.Warning)

            if not self._should_stop:
                self.scan_notification.emit(self.tr("Сканирование завершено."), QSystemTrayIcon.Information)
        except Exception as e:
            self.scan_notification.emit(self.tr("Критическая ошибка сканирования: {0}").format(e), QSystemTrayIcon.Critical)
        finally:
            self.finished.emit()


class CleanupWorker(QObject):
    """
    Рабочий, выполняющий очистку файлов в отдельном потоке.
    """
    finished = Signal()
    progress = Signal(str)
    cleanup_notification = Signal(str, QSystemTrayIcon.MessageIcon)

    def __init__(self, history_manager, watched_items: List[Dict]):
        super().__init__()
        self.history_manager = history_manager
        self.watched_items = watched_items
        self._should_stop = False

    def stop(self):
        self._should_stop = True

    def run(self):
        self.cleanup_notification.emit(
            self.tr("Началась фоновая очистка истории файлов..."),
            QSystemTrayIcon.Information
        )
        messages, files_deleted = self.history_manager.clean_unwatched_files_in_db(
            self.watched_items,
            lambda: self._should_stop
        )
        for msg, icon_type in messages:
            self.cleanup_notification.emit(msg, icon_type)

        if not self._should_stop:
            if files_deleted > 0:
                self.cleanup_notification.emit(self.tr("Фоновая очистка истории завершена. Удалено {0} записей.").format(files_deleted), QSystemTrayIcon.Information)
            else:
                self.cleanup_notification.emit(self.tr("Фоновая очистка истории завершена. Нет файлов для удаления."), QSystemTrayIcon.Information)
        else:
            self.cleanup_notification.emit(self.tr("Фоновая очистка истории прервана пользователем."), QSystemTrayIcon.Warning)

        self.finished.emit()


class HistoryManager(QObject):
    """
    Управляет хранилищем версий файлов.
    """
    scan_started = Signal()
    scan_finished = Signal()
    cleanup_started = Signal()
    cleanup_finished = Signal()
    file_list_updated = Signal()
    version_added = Signal(int)
    history_notification = Signal(str, QSystemTrayIcon.MessageIcon)
    scan_progress = Signal(str)

    # Новый сигнал для обновления информации о хранилище
    # Аргументы: (процент_заполнения_иконки_0_1, размер_хранилища_форматировано,
    #             свободное_место_диска_форматировано, процент_для_тултипа_0_100)
    storage_info_updated = Signal(float, str, str, float)

    DB_NAME = "metadata.db"
    OBJECTS_DIR = "objects"
    STORAGE_SCAN_INTERVAL_MS = 60 * 1000 # 1 минута

    # --- Списки поддерживаемых расширений для предпросмотра ---
    TEXT_EXTENSIONS = {'.txt', '.log', '.md', '.py', '.json', '.xml', '.html', '.css', '.js', '.csv'}
    IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.ico', '.webp', '.tiff', '.tif', '.svg'}
    PDF_EXTENSIONS = {'.pdf'}
    DOCX_EXTENSIONS = {'.docx'}
    XLSX_EXTENSIONS = {'.xlsx'}

    def __init__(self, storage_path: Path, parent=None):
        super().__init__(parent)
        self.storage_path = storage_path
        self.db_path = self.storage_path / self.DB_NAME
        self.objects_path = self.storage_path / self.OBJECTS_DIR
        self._scan_thread = None
        self._scan_worker = None
        self._is_scan_running = False
        self._cleanup_thread = None
        self._cleanup_worker = None
        self._is_cleanup_running = False
        self._db_connection_lock = threading.RLock()
        self._setup_storage()
        self._db_connection = sqlite3.connect(self.db_path, check_same_thread=False)
        self._db_connection.execute("PRAGMA foreign_keys = ON")
        self._setup_database()
        # --- Система асинхронной очереди задач для предотвращения deadlock ---
        self._pending_operation: Optional[str] = None
        self._pending_args: Optional[tuple] = None
        self._temp_preview_files: List[Path] = [] # Список временных файлов для очистки

        # --- Таймер для периодического обновления информации о хранилище ---
        self._storage_info_timer = QTimer(self)
        self._storage_info_timer.setInterval(self.STORAGE_SCAN_INTERVAL_MS)
        self._storage_info_timer.timeout.connect(self.update_storage_info)
        self._storage_info_timer.start()

        # Первоначальное обновление информации о хранилище (выполняется синхронно)
        self.update_storage_info()

    def _calculate_hash(self, file_path: Path) -> str | None:
        try:
            sha256 = hashlib.sha256()
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    sha256.update(chunk)
            return sha256.hexdigest()
        except (OSError, IOError):
            return None

    def _format_size(self, size_bytes: int) -> str:
        """Форматирует размер файла в удобочитаемый вид."""
        if size_bytes < 1024:
            return self.tr("{0} B").format(size_bytes)
        elif size_bytes < 1024 ** 2:
            return self.tr("{0:.1f} KB").format(size_bytes / 1024)
        elif size_bytes < 1024 ** 3:
            return self.tr("{0:.1f} MB").format(size_bytes / (1024 ** 2))
        else:
            return self.tr("{0:.1f} GB").format(size_bytes / (1024 ** 3))

    def _request_stop_all_workers(self):
        """Отправляет неблокирующий запрос на остановку всем активным рабочим."""
        if self._is_scan_running and self._scan_worker:
            self._scan_worker.stop()
        if self._is_cleanup_running and self._cleanup_worker:
            self._cleanup_worker.stop()

    def start_scan(self, items_to_scan: List[Dict]):
        if self._is_scan_running or self._is_cleanup_running:
            self._pending_operation = "scan"
            self._pending_args = (items_to_scan,)
            self._request_stop_all_workers()
            return

        if not items_to_scan:
            self.history_notification.emit(self.tr("Сканирование не начато: нет элементов для сканирования."), QSystemTrayIcon.Information)
            return

        self._is_scan_running = True
        self.scan_started.emit()
        self._scan_thread = QThread(self)
        self._scan_worker = ScanWorker(self, items_to_scan)
        self._scan_worker.moveToThread(self._scan_thread)
        self._scan_thread.started.connect(self._scan_worker.run)
        self._scan_worker.finished.connect(self._on_scan_finished_internal)
        self._scan_worker.scan_notification.connect(self.history_notification)
        self._scan_worker.progress.connect(self.scan_progress)
        self._scan_thread.start()

    def start_cleanup(self, watched_items: List[Dict]):
        if self._is_scan_running or self._is_cleanup_running:
            self._pending_operation = "cleanup"
            self._pending_args = (watched_items,)
            self._request_stop_all_workers()
            return

        self._is_cleanup_running = True
        self.cleanup_started.emit()
        self._cleanup_thread = QThread(self)
        self._cleanup_worker = CleanupWorker(self, watched_items)
        self._cleanup_worker.moveToThread(self._cleanup_thread)
        self._cleanup_thread.started.connect(self._cleanup_worker.run)
        self._cleanup_worker.finished.connect(self._on_cleanup_finished_internal)
        self._cleanup_worker.cleanup_notification.connect(self.history_notification)
        self._cleanup_thread.start()

    @Slot()
    def update_storage_info(self):
        """
        Рассчитывает текущее использование хранилища относительно свободного места на диске
        и отправляет сигнал.
        """
        try:
            # 1. Рассчитываем размер папки хранилища Undoit
            total_undoit_storage_size_bytes = 0
            if self.storage_path.exists():
                for dirpath, dirnames, filenames in os.walk(self.storage_path):
                    for f in filenames:
                        fp = Path(dirpath) / f
                        if fp.is_file() and not fp.is_symlink():
                            total_undoit_storage_size_bytes += fp.stat().st_size

            # 2. Получаем информацию о диске, на котором находится хранилище
            # Передаем корень диска в psutil.disk_usage
            disk_root = str(self.storage_path.anchor if self.storage_path.is_absolute() else self.storage_path.resolve().anchor)
            disk_usage = psutil.disk_usage(disk_root)

            free_disk_space_bytes = disk_usage.free

            # 3. Рассчитываем проценты
            if free_disk_space_bytes > 0:
                # Отношение размера хранилища Undoit к свободному месту на диске
                storage_to_free_ratio = total_undoit_storage_size_bytes / free_disk_space_bytes

                # Процент для иконки (0.0 до 1.0), ограниченный 100%
                icon_fill_percentage = min(storage_to_free_ratio, 1.0)

                # Процент для тултипа (0.0 до 100.0), также ограниченный 100%
                tooltip_percentage = min(storage_to_free_ratio * 100.0, 100.0)
            else:
                # Если свободного места нет, то хранилище занимает 100% (или больше)
                # от доступного свободного места.
                icon_fill_percentage = 1.0 # Полностью заполнено (красный)
                tooltip_percentage = 100.0 # 100%

            # Форматируем размеры
            formatted_storage_size = self._format_size(total_undoit_storage_size_bytes)
            formatted_free_disk_space = self._format_size(free_disk_space_bytes)

            # Отправляем сигнал
            self.storage_info_updated.emit(
                icon_fill_percentage, 
                formatted_storage_size, 
                formatted_free_disk_space, 
                tooltip_percentage
            )

        except (OSError, psutil.Error, FileNotFoundError) as e:
            self.history_notification.emit(
                self.tr("Ошибка при получении информации о хранилище: {0}").format(e),
                QSystemTrayIcon.Warning
            )
            # При ошибке устанавливаем значения по умолчанию
            self.storage_info_updated.emit(0.0, self.tr("Н/Д"), self.tr("Н/Д"), 0.0)


    def _execute_pending_operation(self):
        """Выполняет отложенную операцию после завершения текущей."""
        if not self._pending_operation:
            return

        op = self._pending_operation
        args = self._pending_args
        self._pending_operation = None
        self._pending_args = None

        if op == "scan" and args:
            self.start_scan(*args)
        elif op == "cleanup" and args:
            self.start_cleanup(*args)

    @Slot()
    def _on_scan_finished_internal(self):
        self._is_scan_running = False
        if self._scan_worker:
            self._scan_worker.deleteLater()
        if self._scan_thread:
            self._scan_thread.quit()
            self._scan_thread.wait() # Безопасно ждать здесь, т.к. мы в слоте главного потока
            self._scan_thread.deleteLater()
        self._scan_worker, self._scan_thread = None, None

        # self.file_list_updated.emit() # Теперь обновляется в _add_version_from_path
        self.scan_finished.emit()
        self.update_storage_info() # Обновляем информацию о хранилище после сканирования

        self._execute_pending_operation()

    @Slot()
    def _on_cleanup_finished_internal(self):
        self._is_cleanup_running = False
        if self._cleanup_worker:
            self._cleanup_worker.deleteLater()
        if self._cleanup_thread:
            self._cleanup_thread.quit()
            self._cleanup_thread.wait() # Безопасно ждать здесь
            self._cleanup_thread.deleteLater()
        self._cleanup_worker, self._cleanup_thread = None, None

        self.file_list_updated.emit() # Очистка может удалить файлы, поэтому обновляем список
        self.cleanup_finished.emit()
        self.update_storage_info() # Обновляем информацию о хранилище после очистки

        self._execute_pending_operation()

    def add_initial_version(self, file_path: Path):
        with self._db_connection_lock:
            cursor = self._db_connection.cursor()
            cursor.execute("SELECT id FROM tracked_files WHERE original_path = ?", (str(file_path),))
            if cursor.fetchone():
                return
            # _add_version_from_path теперь сам испускает сигналы, если файл новый
            self._add_version_from_path(file_path)

    @Slot(str)
    def add_file_version(self, file_path_str: str):
        file_path = Path(file_path_str)
        if not file_path.is_file(): return

        file_hash = self._calculate_hash(file_path)
        if not file_hash:
            self.history_notification.emit(self.tr("Ошибка: не удалось рассчитать хеш для файла {0}").format(file_path.name), QSystemTrayIcon.Warning)
            return

        was_new_file, file_id, error_message = False, -1, None
        with self._db_connection_lock:
            cursor = self._db_connection.cursor()
            cursor.execute("SELECT v.sha256_hash FROM versions v JOIN tracked_files tf ON v.file_id = tf.id WHERE tf.original_path = ? ORDER BY v.timestamp DESC LIMIT 1", (str(file_path),))
            last_version = cursor.fetchone()
            if last_version and last_version[0] == file_hash: return

            result_tuple = self._add_version_from_path(file_path, file_hash)
            if result_tuple:
                # _add_version_from_path уже испустил file_list_updated если was_new_file
                was_new_file, file_id = result_tuple
            else:
                error_message = self.tr("Не удалось сохранить версию файла {0}.").format(file_path.name)

        if error_message:
            self.history_notification.emit(error_message, QSystemTrayIcon.Critical)
            return

        # version_added всегда испускается, если версия добавлена (новый или старый файл)
        self.version_added.emit(file_id)
        if was_new_file:
            # self.file_list_updated.emit() # Теперь испускается в _add_version_from_path
            self.history_notification.emit(self.tr("Добавлен новый файл для отслеживания: {0}").format(file_path.name), QSystemTrayIcon.Information)
        else:
            self.history_notification.emit(self.tr("Сохранена новая версия файла: {0}").format(file_path.name), QSystemTrayIcon.Information)

        self.update_storage_info() # Обновляем информацию о хранилище после добавления версии

    def get_all_tracked_files(self) -> List[tuple]:
        with self._db_connection_lock:
            cursor = self._db_connection.cursor()
            cursor.execute("SELECT id, original_path FROM tracked_files ORDER BY original_path ASC")
            return cursor.fetchall()

    def get_versions_for_file(self, file_id: int) -> List[tuple]:
        with self._db_connection_lock:
            cursor = self._db_connection.cursor()
            cursor.execute("SELECT timestamp, sha256_hash, file_size FROM versions WHERE file_id = ? ORDER BY timestamp DESC", (file_id,))
            return cursor.fetchall()

    def get_object_path(self, sha256_hash: str) -> Path | None:
        object_path = self.objects_path / sha256_hash[:2] / sha256_hash[2:]
        return object_path if object_path.exists() else None

    def _add_temp_preview_file(self, temp_file_path: Path):
        """Добавляет временный файл предпросмотра в список для последующей очистки."""
        self._temp_preview_files.append(temp_file_path)

    def _cleanup_temp_preview_files(self):
        """Удаляет все временные файлы предпросмотра."""
        for temp_file in self._temp_preview_files:
            try:
                if temp_file.exists():
                    os.remove(temp_file)
            except OSError as e:
                # Можно добавить логирование, но уведомление пользователя здесь не нужно
                pass
        self._temp_preview_files.clear()


    def get_file_content_for_preview(self, object_path: Path, original_file_extension: str) -> Tuple[str, Optional[str]]:
        """
        Извлекает содержимое файла для предпросмотра.
        Возвращает кортеж (тип_контента, данные_контента).
        Тип контента: "text", "image", "error", "unsupported".
        Данные контента: строка для текста, путь к временному изображению.
        """
        file_extension = original_file_extension.lower()
        extracted_content = None

        try:
            if file_extension in self.TEXT_EXTENSIONS:
                with open(object_path, 'r', encoding='utf-8', errors='replace') as f:
                    extracted_content = f.read(1024 * 10)
                if extracted_content:
                    return "text", extracted_content
                else:
                    return "unsupported", self.tr("Файл пуст или текст не найден.")

            elif file_extension in self.IMAGE_EXTENSIONS:
                return "image", str(object_path)

            elif file_extension in self.PDF_EXTENSIONS:
                try:
                    doc = fitz.open(object_path)

                    # --- Попытка извлечь текст ---
                    pdf_text_content = ""
                    for page_num in range(doc.page_count):
                        page = doc[page_num]
                        pdf_text_content += page.get_text()
                        if len(pdf_text_content) > 1024 * 10: # Ограничиваем размер извлеченного текста
                            pdf_text_content = pdf_text_content[:1024 * 10] + "\n..."
                            break

                    if pdf_text_content.strip(): # Если текст найден и он не пустой
                        doc.close()
                        return "text", pdf_text_content

                    # --- Если текст не найден (например, PDF только из изображений), рендерим первую страницу ---
                    self.history_notification.emit(
                        self.tr("Текст не найден в PDF файле. Попытка отобразить первую страницу как изображение."),
                        QSystemTrayIcon.Information
                    )
                    page = doc[0] # Берем первую страницу
                    pix = page.get_pixmap(matrix=fitz.Matrix(150/72, 150/72)) # Рендерим с 150 DPI

                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        temp_img_path = Path(tmp.name)
                    pix.save(str(temp_img_path))

                    doc.close()
                    self._add_temp_preview_file(temp_img_path)

                    return "image", str(temp_img_path) 

                except Exception as e:
                    if 'doc' in locals() and not doc.is_closed:
                        doc.close()
                    return "text", self.tr("Не удалось открыть PDF-файл для предпросмотра: {0}. Возможно, он поврежден или защищен.").format(e)

            elif file_extension in self.DOCX_EXTENSIONS:
                try:
                    document = Document(object_path)
                    full_text = [para.text for para in document.paragraphs if para.text.strip()] # Собираем только непустые параграфы
                    extracted_content = "\n".join(full_text[:50]) + ("\n..." if len(full_text) > 50 else "")

                    if extracted_content.strip():
                        return "text", extracted_content
                    else:
                        return "unsupported", self.tr("Текст не найден в DOCX файле.")
                except Exception as e:
                    return "text", self.tr("Не удалось открыть DOCX-файл для предпросмотра: {0}. Возможно, он поврежден.").format(e)

            elif file_extension in self.XLSX_EXTENSIONS:
                try:
                    workbook = load_workbook(object_path, read_only=True, data_only=True)
                    sheet_names = workbook.sheetnames

                    preview_lines = []
                    if sheet_names:
                        preview_lines.append(f"Листы: {', '.join(sheet_names)}")

                        sheet = workbook[sheet_names[0]]
                        preview_lines.append(f"Первые 10 строк листа '{sheet_names[0]}':")
                        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10)):
                            row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                            preview_lines.append(f"{', '.join(row_values)}")

                    extracted_content = "\n".join(preview_lines)

                    if extracted_content.strip():
                        return "text", extracted_content
                    else:
                        return "unsupported", self.tr("Данные не найдены в XLSX файле.")
                except Exception as e:
                    return "text", self.tr("Не удалось открыть XLSX-файл для предпросмотра: {0}. Возможно, он поврежден.").format(e)

            else:
                return "unsupported", self.tr("Предпросмотр недоступен для этого типа файла.")

        except Exception as e:
            return "error", self.tr("Критическая ошибка при предпросмотре файла: {0}").format(e)


    def _add_version_from_path(self, file_path: Path, precalculated_hash: str = None) -> Optional[Tuple[bool, int]]:
        file_hash = precalculated_hash or self._calculate_hash(file_path)
        if not file_hash: return None
        try:
            file_size = file_path.stat().st_size
        except (FileNotFoundError, OSError):
            return None
        object_subdir = self.objects_path / file_hash[:2]
        object_path = object_subdir / file_hash[2:]
        if not object_path.exists():
            try:
                object_subdir.mkdir(exist_ok=True)
                shutil.copy2(file_path, object_path)
            except IOError:
                return None
        cursor = self._db_connection.cursor()
        cursor.execute("SELECT id FROM tracked_files WHERE original_path = ?", (str(file_path),))
        file_id_result = cursor.fetchone()
        was_new_file = not file_id_result
        if was_new_file:
            cursor.execute("INSERT INTO tracked_files (original_path) VALUES (?)", (str(file_path),))
            file_id = cursor.lastrowid
        else:
            file_id = file_id_result[0]
        timestamp = datetime.now().isoformat()
        cursor.execute("INSERT INTO versions (file_id, timestamp, sha256_hash, file_size) VALUES (?, ?, ?, ?)", (file_id, timestamp, file_hash, file_size))
        try:
            self._db_connection.commit()
            if was_new_file:
                self.version_added.emit(file_id)
                self.file_list_updated.emit()
            return was_new_file, file_id
        except sqlite3.Error:
            self._db_connection.rollback()
            return None

    def clean_unwatched_files_in_db(self, watched_items: List[Dict], should_stop_callback=None) -> Tuple[List, int]:
        messages, files_deleted_count, file_ids_to_delete = [], 0, []
        watched_files = {Path(item['path']).resolve() for item in watched_items if item['type'] == 'file'}
        watched_folders = [
            (Path(item['path']).resolve(), {Path(ex).resolve() for ex in item.get('exclusions', [])})
            for item in watched_items if item['type'] == 'folder'
        ]
        with self._db_connection_lock:
            cursor = self._db_connection.cursor()
            cursor.execute("SELECT id, original_path FROM tracked_files")
            all_tracked_files = cursor.fetchall()
            for file_id, original_path_str in all_tracked_files:
                if should_stop_callback and should_stop_callback():
                    messages.append((self.tr("Очистка прервана пользователем."), QSystemTrayIcon.Warning))
                    return messages, files_deleted_count
                try:
                    original_path = Path(original_path_str).resolve()
                except (OSError, RuntimeError): continue # Пропускаем пути, которые больше не существуют
                is_watched = False
                if original_path in watched_files:
                    is_watched = True
                else:
                    for folder_path, exclusion_paths in watched_folders:
                        if original_path.is_relative_to(folder_path):
                            is_excluded = any(original_path.is_relative_to(ex_path) for ex_path in exclusion_paths)
                            if not is_excluded:
                                is_watched = True
                                break
                if not is_watched:
                    file_ids_to_delete.append(file_id)
            if file_ids_to_delete:
                try:
                    placeholders = ','.join('?' for _ in file_ids_to_delete)
                    cursor.execute(f"DELETE FROM versions WHERE file_id IN ({placeholders})", file_ids_to_delete)
                    cursor.execute(f"DELETE FROM tracked_files WHERE id IN ({placeholders})", file_ids_to_delete)
                    self._db_connection.commit()
                    files_deleted_count = len(file_ids_to_delete)
                    messages.append((self.tr("Удалено {0} записей о файлах, которые больше не отслеживаются.").format(files_deleted_count), QSystemTrayIcon.Information))
                except sqlite3.Error as e:
                    messages.append((self.tr("Ошибка при удалении записей из БД: {0}").format(e), QSystemTrayIcon.Critical))
                    self._db_connection.rollback()
        return messages, files_deleted_count

    def _setup_storage(self):
        try:
            self.storage_path.mkdir(parents=True, exist_ok=True)
            self.objects_path.mkdir(exist_ok=True)
        except OSError as e:
            self.history_notification.emit(self.tr("Ошибка создания папки хранилища: {0}").format(e), QSystemTrayIcon.Critical)

    def _setup_database(self):
        with self._db_connection_lock:
            try:
                cursor = self._db_connection.cursor()
                cursor.execute("CREATE TABLE IF NOT EXISTS tracked_files (id INTEGER PRIMARY KEY, original_path TEXT NOT NULL UNIQUE)")
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS versions (
                        id INTEGER PRIMARY KEY, file_id INTEGER NOT NULL, timestamp TEXT NOT NULL,
                        sha256_hash TEXT NOT NULL, file_size INTEGER NOT NULL,
                        FOREIGN KEY (file_id) REFERENCES tracked_files (id) ON DELETE CASCADE
                    )""")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_tracked_files_path ON tracked_files (original_path)")
                cursor.execute("CREATE INDEX IF NOT EXISTS idx_versions_file_id ON versions (file_id)")
                self._db_connection.commit()
            except sqlite3.Error as e:
                self.history_notification.emit(self.tr("Ошибка инициализации базы данных: {0}").format(e), QSystemTrayIcon.Critical)

    def close(self):
        self._request_stop_all_workers()
        if self._scan_thread and self._scan_thread.isRunning():
            self._scan_thread.wait(500)
        if self._cleanup_thread and self._cleanup_thread.isRunning():
            self._cleanup_thread.wait(500)

        self._storage_info_timer.stop() # Останавливаем таймер

        if self._db_connection:
            with self._db_connection_lock:
                self._db_connection.close()

        self._cleanup_temp_preview_files()
